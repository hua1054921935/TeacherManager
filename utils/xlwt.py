# coding=utf-8
# 新建excil
import openpyxl
from openpyxl import load_workbook

def create_excel(wbname):
    wb=openpyxl.Workbook()
    wbname=wbname+'.xlsx'
    wb.save(filename=wbname)
    print(wbname+'建立完成')
    return wbname

def save_data_excel(data,fields,shellname,wbname):
    print('写入表格')
    newwbname=create_excel(wbname)
    wb=openpyxl.load_workbook(filename=newwbname)
    sheel=wb.active
    sheel.title=shellname

    field=1
    for field in range(1,len(fields)+1):
        _ = sheel.cell(row=1, column=field, value=str(fields[field - 1]))

    row1 = 1
    col1 = 0
    for row1 in range(2, len(data) + 2):  # 写入数据
        for col1 in range(1, len(data[row1 - 2]) + 1):
            _ = sheel.cell(row=row1, column=col1, value=str(data[row1 - 2][col1 - 1]))

    wb.save(filename=newwbname)
    print("保存成功")
    return newwbname

def read_xls(self,request, obj, change):
    print(obj)
    wb = load_workbook(filename='../File/'+obj+'.xlsx')
    ws = wb.active

    # rows = table.max_row  # 获取行数
    # cols = table.max_column # 获取列数
    # Data = table.cell(row=rows, column=cols).value
    print(ws)



if __name__ == '__main__':
        wbname='新建'
        fields=['姓名','性别','年龄']
        data=[['1','2','3']]
        shellname='表格1'
        save_data_excel(data,fields,shellname,wbname)
        # readexl()

